from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from openpyxl import load_workbook
from datetime import datetime
import os
import json

from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
from io import BytesIO

app = Flask(__name__)
app.secret_key = 'your_secret_key'


# ID file Google Drive
FILE_ID = '1qTz1QNwRu3wvTlM_EL43FvnU_dKQ6Qpc'  # Ganti dengan file ID dari file Excel di Google Drive
PENGUNJUNG_SHEET = 'Pengunjung' #Nama sheet di file excel yang digunakan untuk menyimpan data visitor
KUNJUNGAN_SHEET = 'Kunjungan' #Nama sheet di file excel yang digunakan untuk merekam data kunjungan visitor


#Konfigurasi Google Drive
# Konfigurasi Google Drive dengan error handling
def initialize_drive():
    try:
        # Muat kredensial dari environment variable
        google_credentials = json.loads(os.environ.get('GOOGLE_CREDENTIALS', '{}'))
        
        # Gunakan temporary directory yang aman
        temp_cred_path = '/tmp/credentials.json'
        
        # Simpan kredensial sementara
        with open(temp_cred_path, 'w') as cred_file:
            json.dump(google_credentials, cred_file)
        
        # Autentikasi
        gauth = GoogleAuth()
        gauth.LoadCredentialsFile(temp_cred_path)
        
        if not gauth.credentials or gauth.credentials.invalid:
            raise Exception("Invalid Google Drive credentials")
            
        return GoogleDrive(gauth)
    except Exception as e:
        print(f"Error initializing Google Drive: {str(e)}")
        return None

def load_excel_file(drive):
    """Download dan baca file Excel dari Google Drive dengan error handling."""
    try:
        if not drive:
            raise Exception("Google Drive not initialized")
            
        file = drive.CreateFile({'id': FILE_ID})
        temp_path = '/tmp/temp.xlsx'
        file.GetContentFile(temp_path)
        workbook = load_workbook(temp_path)
        
        # Cleanup
        if os.path.exists(temp_path):
            os.remove(temp_path)
            
        return workbook
    except Exception as e:
        print(f"Error loading Excel file: {str(e)}")
        return None

def save_excel_file(drive, workbook):
    """Simpan file Excel ke Google Drive dengan error handling."""
    try:
        if not drive:
            raise Exception("Google Drive not initialized")
            
        with BytesIO() as output:
            workbook.save(output)
            output.seek(0)
            file = drive.CreateFile({'id': FILE_ID})
            file.SetContentString(output.getvalue().decode('latin1'))
            file.Upload()
        return True
    except Exception as e:
        print(f"Error saving Excel file: {str(e)}")
        return False

@app.route("/")
def index():
    return redirect(url_for("home"))

@app.route("/visilog", methods=["POST", "GET"])
def home():
    if request.method == 'POST':

        try:
            id_pengunjung = request.form.get('id_pengunjung')

            # Initialize Google Drive
            drive = initialize_drive()
            if not drive:
                return jsonify({
                    'success': False,
                    'message': 'Failed to initialize Google Drive',
                    'redirect': url_for('home')
                }), 500
            
            # Load workbook
            wb = load_excel_file(drive)
            if not wb:
                return jsonify({
                    'success': False,
                    'message': 'Failed to load Excel file',
                    'redirect': url_for('home')
                }), 500
                
            pengunjung_sheet = wb[PENGUNJUNG_SHEET]
            kunjungan_sheet = wb[KUNJUNGAN_SHEET]
            
            # Cari data pengunjung
            pengunjung = None
            for row in pengunjung_sheet.iter_rows(min_row=2, values_only=True):
                if str(row[0]) == id_pengunjung:
                    pengunjung = row
                    break

            if pengunjung:
                # Rekam kunjungan
                TANGGAL_KUNJUNGAN = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                kunjungan_sheet.append([pengunjung[0], pengunjung[1], pengunjung[2], TANGGAL_KUNJUNGAN])
                save_excel_file(wb)  # Simpan workbook kembali ke Google Drive

                return jsonify({
                    'success': True,
                    'nama': pengunjung[1],
                    'redirect': url_for('home')
                })

            else:
                return jsonify({
                    'success': False,
                    'redirect': url_for('home')
                })
            
        except Exception as e:
            flash(f"Terjadi kesalahan: {e}", 'danger')

        return redirect(url_for('home'))

    return render_template('visitor-log.html')



# Konfigurasi untuk production
app.debug = False if os.environ.get('FLASK_ENV') == 'production' else True

#if __name__ == '__main__':
    #app.run(debug = True, host = '0.0.0.0', port = 5555)
